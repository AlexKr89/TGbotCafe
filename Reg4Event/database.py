# database.py
import openpyxl

class Database:
    def __init__(self, filename):
        self.filename = filename

    def get_events(self):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.active
        events = [(row[0], row[1], row[2]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] and row[1] and row[2]]
        wb.close()
        return events

class RegistrationDatabase:
    def __init__(self, filename):
        self.filename = filename

    def save_registration(self, event_name, user_info):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.active

        # Find the row where the event_name matches
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0] == event_name:
                # Append user_info and current timestamp to the row
                registration_data = [user_info, datetime.now()]
                sheet.append([event_name, *registration_data])
                break

        wb.save(self.filename)
        wb.close()
